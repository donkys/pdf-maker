import uuid
import openpyxl
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.colors import red
from cryptography.fernet import Fernet
import os
from io import BytesIO

# Paths declarations
SECRET_KEY_PATH = "./asset/secret.key"
FONT_PATH = './font/Kanit-Regular.ttf'
INPUT_PDF_PATH = "./original/original.pdf"
INPUT_XLSX_PATH = "input.xlsx"
OUTPUT_XLSX_PATH = "./asset/output.xlsx"
SEQUENCE_FILE = "sequence.txt"
OUTPUT_DIR = "output"

def load_key():
    return open(SECRET_KEY_PATH, "rb").read()

def encrypt_message(message, key):
    f = Fernet(key)
    encrypted_message = f.encrypt(message.encode())
    return encrypted_message.decode()

def overlay_text_to_pdf(input_pdf_path, output_pdf_path, name, facebook):
    packet = BytesIO()
    can = canvas.Canvas(packet, pagesize=(595.27,841.89))  # A4

    # Register the font
    font_name = "Kanit"
    pdfmetrics.registerFont(TTFont(font_name, FONT_PATH))

    can.setFont(font_name, 7)
    can.setFillColor(red)
    text = "ไฟล์นี้เป็นลิขสิทธิ์ของเพื่อนวิชาการและไฟล์นี้ถูกซื้อโดยคุณ " + str(name) + " - " + str(facebook) + " ห้ามเผยแพร่หรือแจกผู้ซื้อมีสิทธิ์สามารถอ่านได้ผู้เดียว."
    can.drawString(10, 10, text)

    can.save()
    packet.seek(0)
    new_pdf = PdfReader(packet)

    existing_pdf = PdfReader(open(input_pdf_path, "rb"))
    output = PdfWriter()
    for i in range(len(existing_pdf.pages)):
        page = existing_pdf.pages[i]
        page.merge_page(new_pdf.pages[0])
        output.add_page(page)

    with open(output_pdf_path, "wb") as f:
        output.write(f)

def modify_pdf(input_pdf_path, output_pdf_path, name, facebook, key):
    overlay_text_to_pdf(input_pdf_path, output_pdf_path, name, facebook)
    reader = PdfReader(output_pdf_path)
    writer = PdfWriter()
    for page_num in range(len(reader.pages)):
        page = reader.pages[page_num]
        writer.add_page(page)

    unique_id_raw = str(uuid.uuid4())
    unique_id_encrypted = encrypt_message(unique_id_raw, key)

    metadata = reader.metadata
    writer.add_metadata(metadata)
    writer.add_metadata({'/UniqueID': unique_id_encrypted})

    with open(output_pdf_path, 'wb') as out:
        writer.write(out)

    return unique_id_encrypted

def read_sequence_from_file(filename=SEQUENCE_FILE):
    if not os.path.exists(filename):
        return 1
    with open(filename, "r") as f:
        return int(f.read().strip())

def write_sequence_to_file(sequence, filename=SEQUENCE_FILE):
    with open(filename, "w") as f:
        f.write(str(sequence))

def append_to_excel(filename, new_rows):
    """
    Append new rows to an existing Excel file
    """
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Append the new rows
    for row in new_rows:
        sheet.append(row)

    # Save the workbook
    workbook.save(filename)

key = load_key()
input_rows = []
output_rows = []

# Check and make directory if it doesn't exist
if not os.path.exists(OUTPUT_DIR):
    os.mkdir(OUTPUT_DIR)

print("Processing input.xlsx...")
counter = read_sequence_from_file()

workbook = openpyxl.load_workbook(INPUT_XLSX_PATH)
sheet = workbook.active
rows_raw = list(sheet.iter_rows(values_only=True))

# Convert all values to string
rows = [[str(cell) if cell is not None else None for cell in row] for row in rows_raw]

for row in rows[1:]:
    name = row[0]
    facebook = row[1]
    print(f"Processing: {name}")

    output_pdf_path = os.path.join(OUTPUT_DIR, f"{counter:04}_{name}.pdf")
    try:
        input_pdf_path = INPUT_PDF_PATH
        unique_id = modify_pdf(input_pdf_path, output_pdf_path, name, facebook, key)
        output_rows.append([name, facebook, unique_id])
        counter += 1
    except Exception as e:
        print(f"Error processing {name}: {e}")
        input_rows.append([name, facebook])

write_sequence_to_file(counter)

print("Updating input.xlsx and writing to output.xlsx...")

input_workbook = openpyxl.Workbook()
input_sheet = input_workbook.active
input_sheet.append(["ชื่อ-สกุล", "Facebook"])
for row in input_rows:
    input_sheet.append(row)
input_workbook.save(INPUT_XLSX_PATH)

# Append to output.xlsx instead of overwriting it
if not os.path.exists(OUTPUT_XLSX_PATH):
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    output_sheet.append(["ชื่อ-สกุล", "Facebook", "Unique ID"])
    output_workbook.save(OUTPUT_XLSX_PATH)

append_to_excel(OUTPUT_XLSX_PATH, output_rows)

print("Done!")
